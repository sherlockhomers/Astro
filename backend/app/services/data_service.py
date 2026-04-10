from __future__ import annotations

import csv
import hashlib
import re
import sqlite3
from pathlib import Path
from typing import Any
from urllib.parse import quote

from app.config import settings
from app.services.sqlite_service import get_sqlite_connection


TOKEN_RE = re.compile(r"[A-Za-z0-9\u4e00-\u9fff]+")
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
EXCEL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TEXT_EXTS = {".txt", ".md", ".markdown", ".jsonl", ".json", ".pdf"}


class DataService:
    def __init__(self) -> None:
        self._entities: list[dict[str, Any]] = []
        self._extra_entities: list[dict[str, Any]] = []
        self._images: dict[str, dict[str, Any]] = {}
        self._source_root: str | None = None
        self._loaded = False
        self._revision = 0
        self._db_path = settings.sqlite_path
        self._init_db()
        self._extra_entities = self._load_persisted_extra_entities()

    @property
    def loaded(self) -> bool:
        return self._loaded

    @property
    def entity_count(self) -> int:
        return len(self._entities)

    @property
    def source_root(self) -> str | None:
        return self._source_root

    @property
    def revision(self) -> int:
        return self._revision

    def scan_csv_root(self, csv_root: str) -> list[dict[str, Any]]:
        # 向后兼容旧接口名；实际已支持 CSV 目录或 Excel 文件。
        return self.scan_data_source(csv_root)

    def scan_data_source(self, data_source: str) -> list[dict[str, Any]]:
        source = Path(data_source)
        if not source.exists():
            raise ValueError(f"数据源不存在: {data_source}")
        if source.is_dir():
            return self._scan_csv_directory(source)
        if source.is_file() and source.suffix.lower() in EXCEL_EXTS:
            return self._scan_excel_file(source)
        raise ValueError(f"数据源必须是 CSV 目录或 Excel 文件(.xlsx): {data_source}")

    def load_csv_root(self, csv_root: str, categories: list[str] | None = None) -> dict[str, Any]:
        # 向后兼容旧接口名；实际已支持 CSV 目录或 Excel 文件。
        return self.load_data_source(csv_root, categories)

    def load_data_source(self, data_source: str, categories: list[str] | None = None) -> dict[str, Any]:
        source = Path(data_source)
        if not source.exists():
            raise ValueError(f"数据源不存在: {data_source}")
        # Make local image path resolution available during current load loop.
        self._source_root = str(source)

        cat_filter = {c.strip().lower() for c in (categories or []) if c.strip()}
        entities: list[dict[str, Any]] = []
        images: dict[str, dict[str, Any]] = {}
        file_count = 0

        if source.is_dir():
            files = sorted(source.glob("*.csv"))
            for file in files:
                category = file.stem.lower()
                if cat_filter and category not in cat_filter:
                    continue
                file_count += 1
                with file.open("r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    for idx, row in enumerate(reader, start=1):
                        normalized = self._normalize_row(row)
                        entity = self._build_entity(
                            normalized=normalized,
                            fallback_category=category,
                            source_file=file.name,
                            row_index=idx,
                        )
                        image_id = self._register_image(normalized, entity["name"], file.name, images)
                        if image_id:
                            entity["image_id"] = image_id
                        entities.append(entity)
        elif source.is_file() and source.suffix.lower() in EXCEL_EXTS:
            wb = self._open_workbook(source)
            try:
                for sheet in wb.worksheets:
                    sheet_name = str(sheet.title).strip() or "sheet"
                    sheet_category = sheet_name.lower()
                    if cat_filter and sheet_category not in cat_filter:
                        continue
                    file_count += 1
                    headers = self._sheet_headers(sheet)
                    if not headers:
                        continue
                    for idx, raw_row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                        if self._row_is_empty(raw_row):
                            continue
                        row_map = {headers[i]: raw_row[i] for i in range(min(len(headers), len(raw_row)))}
                        normalized = self._normalize_row(row_map)
                        entity = self._build_entity(
                            normalized=normalized,
                            fallback_category=sheet_category,
                            source_file=f"{source.name}:{sheet_name}",
                            row_index=idx,
                        )
                        image_id = self._register_image(
                            normalized,
                            entity["name"],
                            f"{source.name}:{sheet_name}",
                            images,
                        )
                        if image_id:
                            entity["image_id"] = image_id
                        entities.append(entity)
            finally:
                wb.close()
            sidecar = source.parent / "images_catalog.csv"
            if settings.auto_load_images_catalog and sidecar.exists() and sidecar.is_file():
                sidecar_category = "images_catalog"
                if not cat_filter or sidecar_category in cat_filter:
                    file_count += 1
                    with sidecar.open("r", encoding="utf-8", newline="") as f:
                        reader = csv.DictReader(f)
                        for idx, row in enumerate(reader, start=1):
                            normalized = self._normalize_row(row)
                            entity = self._build_entity(
                                normalized=normalized,
                                fallback_category=sidecar_category,
                                source_file=sidecar.name,
                                row_index=idx,
                            )
                            image_id = self._register_image(normalized, entity["name"], sidecar.name, images)
                            if image_id:
                                entity["image_id"] = image_id
                            entities.append(entity)
        else:
            raise ValueError(f"数据源必须是 CSV 目录或 Excel 文件(.xlsx): {data_source}")

        merged_entities = self._merge_entities(entities, self._extra_entities)
        self._entities = merged_entities
        self._images = images
        self._loaded = True
        self._revision += 1
        self._persist_images_metadata(images)

        categories_count = len({e["category"] for e in merged_entities})
        return {
            "loaded": self._loaded,
            "source_root": self._source_root,
            "entity_count": len(merged_entities),
            "category_count": categories_count,
            "file_count": file_count,
            "extra_entity_count": len(self._extra_entities),
        }

    def ingest_text_corpus(
        self,
        text_root: str,
        category_prefix: str = "text_knowledge",
        chunk_size: int = 900,
        overlap: int = 150,
    ) -> dict[str, Any]:
        root = Path(text_root)
        if not root.exists():
            raise ValueError(f"文本语料路径不存在: {text_root}")

        chunk_size = max(300, min(int(chunk_size), 4000))
        overlap = max(0, min(int(overlap), chunk_size // 2))
        category = (category_prefix or "text_knowledge").strip().lower()

        files = self._collect_text_files(root)
        if not files:
            raise ValueError("未发现可导入文本文件（支持 .txt/.md/.jsonl/.json）")

        extra_entities: list[dict[str, Any]] = []
        ingested_files = 0
        chunk_total = 0
        for file_path in files:
            blocks = self._read_text_blocks(file_path)
            if not blocks:
                continue
            ingested_files += 1
            for idx, block in enumerate(blocks, start=1):
                title = str(block.get("title", "")).strip() or file_path.stem
                text = str(block.get("text", "")).strip()
                if not text:
                    continue
                for chunk_idx, chunk in enumerate(self._chunk_text(text, chunk_size, overlap), start=1):
                    if root.is_dir():
                        rel = file_path.relative_to(root).as_posix()
                    else:
                        rel = file_path.name
                    eid_seed = f"{file_path.resolve()}::{idx}::{chunk_idx}::{title}"
                    eid = f"text_{hashlib.sha1(eid_seed.encode('utf-8')).hexdigest()[:20]}"
                    source_file = f"text_corpus/{rel}"
                    extra_entities.append(
                        {
                            "id": eid,
                            "name": title,
                            "description": chunk,
                            "category": category,
                            "source_file": source_file,
                            "raw": {
                                "title": title,
                                "text": chunk,
                                "text_source": str(file_path),
                                "text_chunk_index": chunk_idx,
                                "text_block_index": idx,
                                "text_corpus": True,
                            },
                        }
                    )
                    chunk_total += 1

        if not extra_entities:
            raise ValueError("文本语料中没有可导入的有效内容")

        self._extra_entities = self._merge_entities(self._extra_entities, extra_entities)
        self._persist_extra_entities(extra_entities)
        self._entities = self._merge_entities(self._entities, extra_entities)
        self._loaded = bool(self._entities)
        self._revision += 1

        return {
            "ok": True,
            "text_root": str(root),
            "files": ingested_files,
            "chunks": chunk_total,
            "category": category,
            "entity_count": len(self._entities),
            "extra_entity_count": len(self._extra_entities),
        }

    def upsert_dynamic_fact(
        self,
        entity_name: str,
        payload: dict[str, Any],
        provider: str = "dynamic",
    ) -> dict[str, Any]:
        name = str(payload.get("name") or entity_name or "").strip()
        if not name:
            raise ValueError("dynamic fact requires a non-empty entity name")

        provider_tag = str(provider or payload.get("provider") or "dynamic").strip().lower()
        entity_id = f"dynamic_{hashlib.sha1(name.lower().encode('utf-8')).hexdigest()[:20]}"
        raw = dict(payload)
        raw["dynamic_fact"] = True
        raw["provider"] = provider_tag

        entity = {
            "id": entity_id,
            "name": name,
            "description": self._dynamic_fact_description(raw),
            "category": "dynamic_fact",
            "source_file": f"dynamic/{provider_tag}",
            "raw": raw,
        }

        before_ids = {str(x.get("id", "")) for x in self._extra_entities}
        upserts: list[dict[str, Any]] = [entity]

        # Build a lightweight host-star node for visualization/path retrieval.
        host_star = str(raw.get("host_star") or "").strip()
        if host_star:
            host_id = f"dynamic_star_{hashlib.sha1(host_star.lower().encode('utf-8')).hexdigest()[:20]}"
            host_exists = any(
                str(x.get("name", "")).strip().lower() == host_star.lower()
                for x in self._entities
            ) or any(
                str(x.get("name", "")).strip().lower() == host_star.lower()
                for x in self._extra_entities
            )
            if not host_exists:
                upserts.append(
                    {
                        "id": host_id,
                        "name": host_star,
                        "description": f"Host star node generated from dynamic source ({provider_tag})",
                        "category": "dynamic_fact",
                        "source_file": f"dynamic/{provider_tag}",
                        "raw": {
                            "dynamic_fact": True,
                            "provider": provider_tag,
                            "object_type": "star",
                        },
                    }
                )

        self._extra_entities = self._merge_entities(self._extra_entities, upserts)
        self._entities = self._merge_entities(self._entities, upserts)
        self._persist_extra_entities(upserts)
        self._loaded = bool(self._entities)
        self._revision += 1
        return {
            "ok": True,
            "entity_id": entity_id,
            "name": name,
            "inserted": entity_id not in before_ids,
        }

    def clear_text_corpus(self) -> dict[str, Any]:
        if not self._extra_entities:
            return {
                "ok": True,
                "removed": 0,
                "entity_count": len(self._entities),
            }
        text_ids = {
            str(x.get("id", ""))
            for x in self._extra_entities
            if isinstance(x.get("raw"), dict) and bool(x.get("raw", {}).get("text_corpus"))
        }
        if not text_ids:
            return {
                "ok": True,
                "removed": 0,
                "entity_count": len(self._entities),
            }

        before = len(self._entities)
        self._entities = [e for e in self._entities if str(e.get("id", "")) not in text_ids]
        self._extra_entities = [e for e in self._extra_entities if str(e.get("id", "")) not in text_ids]
        removed = before - len(self._entities)
        self._rewrite_persisted_extra_entities(self._extra_entities)
        self._revision += 1
        return {
            "ok": True,
            "removed": removed,
            "entity_count": len(self._entities),
        }

    def _collect_text_files(self, root: Path) -> list[Path]:
        if root.is_file():
            return [root] if root.suffix.lower() in TEXT_EXTS else []
        files = [p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in TEXT_EXTS]
        files.sort()
        return files

    def _read_text_blocks(self, file_path: Path) -> list[dict[str, str]]:
        suffix = file_path.suffix.lower()
        if suffix in {".txt", ".md", ".markdown"}:
            text = file_path.read_text(encoding="utf-8", errors="ignore").strip()
            if not text:
                return []
            return [{"title": self._guess_title_from_text(file_path, text), "text": text}]
        if suffix == ".pdf":
            return self._read_pdf_blocks(file_path)
        if suffix == ".jsonl":
            rows: list[dict[str, str]] = []
            with file_path.open("r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    payload = self._json_load(line)
                    if not isinstance(payload, dict):
                        continue
                    title = str(
                        payload.get("title")
                        or payload.get("name")
                        or payload.get("heading")
                        or file_path.stem
                    ).strip()
                    text = str(
                        payload.get("text")
                        or payload.get("content")
                        or payload.get("body")
                        or payload.get("description")
                        or ""
                    ).strip()
                    if text:
                        rows.append({"title": title, "text": text})
            return rows
        if suffix == ".json":
            raw = file_path.read_text(encoding="utf-8", errors="ignore").strip()
            payload = self._json_load(raw)
            return self._extract_json_docs(payload, fallback_title=file_path.stem)
        return []

    def _read_pdf_blocks(self, file_path: Path) -> list[dict[str, str]]:
        try:
            from pypdf import PdfReader
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"读取 PDF 失败，请先安装 pypdf: {exc}") from exc

        try:
            reader = PdfReader(str(file_path))
        except Exception:
            return []

        doc_title = ""
        try:
            if reader.metadata and reader.metadata.title:
                doc_title = str(reader.metadata.title).strip()
        except Exception:
            doc_title = ""
        if not doc_title:
            doc_title = file_path.stem

        blocks: list[dict[str, str]] = []
        for idx, page in enumerate(reader.pages, start=1):
            try:
                text = (page.extract_text() or "").strip()
            except Exception:
                text = ""
            if not text:
                continue
            blocks.append({"title": f"{doc_title} - page {idx}", "text": text})
        return blocks

    def _extract_json_docs(self, payload: Any, fallback_title: str) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        if isinstance(payload, dict):
            title = str(payload.get("title") or payload.get("name") or fallback_title).strip()
            text = str(
                payload.get("text")
                or payload.get("content")
                or payload.get("body")
                or payload.get("description")
                or ""
            ).strip()
            if text:
                rows.append({"title": title, "text": text})
            if isinstance(payload.get("items"), list):
                for item in payload["items"]:
                    rows.extend(self._extract_json_docs(item, fallback_title=title))
        elif isinstance(payload, list):
            for item in payload:
                rows.extend(self._extract_json_docs(item, fallback_title=fallback_title))
        return rows

    def _guess_title_from_text(self, file_path: Path, text: str) -> str:
        first_line = next((line.strip() for line in text.splitlines() if line.strip()), "")
        first_line = first_line.replace("\ufeff", "").strip()
        if first_line.startswith("#"):
            first_line = first_line.lstrip("#").strip()
        return first_line[:80] if first_line else file_path.stem

    def _chunk_text(self, text: str, chunk_size: int, overlap: int) -> list[str]:
        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        paras = [p.strip() for p in normalized.split("\n\n") if p.strip()]
        if not paras:
            return []

        chunks: list[str] = []
        current = ""
        for para in paras:
            candidate = para if not current else f"{current}\n\n{para}"
            if len(candidate) <= chunk_size:
                current = candidate
                continue
            if current:
                chunks.append(current)
            current = para
            while len(current) > chunk_size:
                chunks.append(current[:chunk_size])
                tail_start = max(0, chunk_size - overlap)
                current = current[tail_start:]
        if current:
            chunks.append(current)
        # 去重与清理超短片段
        cleaned: list[str] = []
        seen: set[str] = set()
        for chunk in chunks:
            c = " ".join(chunk.split())
            if len(c) < 40 or c in seen:
                continue
            seen.add(c)
            cleaned.append(c)
        return cleaned

    def _dynamic_fact_description(self, payload: dict[str, Any]) -> str:
        parts: list[str] = []

        def add(label: str, key: str) -> None:
            value = payload.get(key)
            if value in (None, "", "unknown"):
                return
            parts.append(f"{label}: {value}")

        add("Host star", "host_star")
        add("Discovery year", "discovery_year")
        add("Discovery method", "discovery_method")
        add("Mass(Earth)", "mass_earth")
        add("Radius(Earth)", "radius_earth")
        add("Orbital period(days)", "orbital_period_days")
        add("Distance(pc)", "distance_pc")
        add("Source", "provider")
        if not parts:
            return str(payload.get("name", "dynamic_fact")).strip()
        return "; ".join(parts)

    def _scan_csv_directory(self, root: Path) -> list[dict[str, Any]]:
        csv_files = sorted(root.glob("*.csv"))
        result: list[dict[str, Any]] = []
        for file in csv_files:
            columns: list[str] = []
            rows = 0
            with file.open("r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                if reader.fieldnames:
                    columns = [str(c).strip() for c in reader.fieldnames if str(c).strip()]
                for row in reader:
                    if any(v not in (None, "") for v in row.values()):
                        rows += 1
            result.append({"name": file.name, "rows": rows, "columns": columns})
        return result

    def _scan_excel_file(self, file_path: Path) -> list[dict[str, Any]]:
        wb = self._open_workbook(file_path)
        result: list[dict[str, Any]] = []
        try:
            for sheet in wb.worksheets:
                headers = self._sheet_headers(sheet)
                if not headers:
                    result.append({"name": f"{file_path.name}:{sheet.title}", "rows": 0, "columns": []})
                    continue
                rows = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if self._row_is_empty(row):
                        continue
                    rows += 1
                result.append(
                    {
                        "name": f"{file_path.name}:{sheet.title}",
                        "rows": rows,
                        "columns": headers,
                    }
                )
        finally:
            wb.close()
        sidecar = file_path.parent / "images_catalog.csv"
        if settings.auto_load_images_catalog and sidecar.exists() and sidecar.is_file():
            columns: list[str] = []
            rows = 0
            with sidecar.open("r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                if reader.fieldnames:
                    columns = [str(c).strip() for c in reader.fieldnames if str(c).strip()]
                for row in reader:
                    if any(v not in (None, "") for v in row.values()):
                        rows += 1
            result.append({"name": sidecar.name, "rows": rows, "columns": columns})
        return result

    def _open_workbook(self, file_path: Path) -> Any:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # noqa: BLE001
            raise ValueError(f"读取 Excel 失败，请先安装 openpyxl: {exc}") from exc
        return load_workbook(filename=str(file_path), read_only=True, data_only=True)

    def _sheet_headers(self, sheet: Any) -> list[str]:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers: list[str] = []
        for idx, value in enumerate(header_row, start=1):
            key = str(value).replace("\ufeff", "").strip().lower() if value is not None else ""
            headers.append(key or f"col_{idx}")
        while headers and not headers[-1]:
            headers.pop()
        return headers

    def _row_is_empty(self, row: Any) -> bool:
        return not any(value not in (None, "") for value in row or ())

    def _normalize_row(self, row: dict[str, Any]) -> dict[str, Any]:
        normalized: dict[str, Any] = {}
        for key, value in row.items():
            nk = str(key).replace("\ufeff", "").strip().lower()
            if not nk:
                continue
            if isinstance(value, str):
                normalized[nk] = value.strip()
            else:
                normalized[nk] = value
        return normalized

    def _pick_first(self, raw: dict[str, Any], keys: tuple[str, ...]) -> str:
        for key in keys:
            value = raw.get(key)
            if value in (None, ""):
                continue
            text = str(value).strip()
            if text:
                return text
        return ""

    def _build_entity(
        self,
        normalized: dict[str, Any],
        fallback_category: str,
        source_file: str,
        row_index: int,
    ) -> dict[str, Any]:
        name = self._pick_first(
            normalized,
            (
                "name",
                "title",
                "object_name",
                "celestial_body",
                "名称",
                "天体名称",
                "标题",
            ),
        )
        entity_id = self._pick_first(normalized, ("id", "编号", "序号", "index"))
        description = self._pick_first(
            normalized,
            ("description", "desc", "summary", "body", "content", "text", "简介", "描述", "说明", "正文", "内容", "详情"),
        )
        category = self._pick_first(
            normalized,
            ("category", "type", "分类", "类别"),
        )
        if not name:
            name = f"{fallback_category}_{row_index}"
        if not entity_id:
            entity_id = f"{fallback_category}:{row_index}"
        if not category:
            category = fallback_category
        return {
            "id": entity_id,
            "name": name,
            "description": description,
            "category": category,
            "source_file": source_file,
            "raw": normalized,
        }

    def get_status(self) -> dict[str, Any]:
        return {
            "loaded": self._loaded,
            "source_root": self._source_root,
            "entity_count": len(self._entities),
            "category_count": len({e["category"] for e in self._entities}),
            "image_count": len(self._images),
            "extra_entity_count": len(self._extra_entities),
            "revision": self._revision,
        }

    def export_entities(self) -> list[dict[str, Any]]:
        return list(self._entities)

    def list_images(self, query: str = "", page: int = 1, page_size: int = 30) -> dict[str, Any]:
        page = max(1, page)
        page_size = max(1, min(page_size, 100))
        needle = query.strip().lower()
        rows = list(self._images.values())
        if not rows:
            rows = self._list_images_from_db()
        if needle:
            rows = [
                row
                for row in rows
                if needle in str(row.get("title", "")).lower()
                or needle in str(row.get("source", "")).lower()
                or needle in str(row.get("image_id", "")).lower()
            ]
        total = len(rows)
        start = (page - 1) * page_size
        end = start + page_size
        items = rows[start:end]
        return {
            "items": items,
            "page": page,
            "page_size": page_size,
            "total": total,
            "has_next": end < total,
        }

    def _list_images_from_db(self) -> list[dict[str, Any]]:
        conn = self._connect()
        try:
            rows = conn.execute(
                """
                SELECT image_id, title, source, ref, kind, url, object_keys_json, bucket
                FROM image_assets
                ORDER BY image_id DESC
                """
            ).fetchall()
            payloads: list[dict[str, Any]] = []
            for row in rows:
                item = dict(row)
                item["object_keys"] = self._json_load(item.pop("object_keys_json", "{}"))
                payloads.append(item)
            return payloads
        finally:
            conn.close()

    def get_image_meta(self, image_id: str) -> dict[str, Any] | None:
        item = self._images.get(image_id)
        if item is not None:
            return item
        conn = self._connect()
        try:
            row = conn.execute(
                """
                SELECT image_id, title, source, ref, kind, url, object_keys_json, bucket
                FROM image_assets
                WHERE image_id = ?
                """,
                (image_id,),
            ).fetchone()
            if row is None:
                return None
            payload = dict(row)
            payload["object_keys"] = self._json_load(payload.pop("object_keys_json", "{}"))
            return payload
        finally:
            conn.close()

    def search(self, query: str, top_k: int) -> list[dict[str, Any]]:
        query_tokens = self._tokenize(query)
        scored: list[tuple[float, dict[str, Any]]] = []

        for entity in self._entities:
            title_text = f"{entity['name']} {entity['category']}"
            content_text = f"{entity['name']} {entity['description']} {entity['raw']}"
            score = self._score(query_tokens, content_text)
            if query.lower() in title_text.lower():
                score += 0.8
            if score <= 0:
                continue
            scored.append((score, entity))

        scored.sort(key=lambda x: x[0], reverse=True)
        items: list[dict[str, Any]] = []
        for score, entity in scored[:top_k]:
            snippet = entity["description"][:140] if entity["description"] else "无描述信息"
            image_url = self._build_image_url(entity.get("raw", {}), entity.get("name", ""), entity.get("source_file", ""))
            items.append(
                {
                    "id": entity["id"],
                    "title": entity["name"],
                    "score": round(float(score), 4),
                    "source": entity["source_file"],
                    "snippet": snippet,
                    "image_url": image_url,
                }
            )
        return items

    def find_best_entity_for_question(self, question: str) -> dict[str, Any] | None:
        lowered = question.lower()
        for entity in self._entities:
            name = str(entity["name"]).strip()
            if not name:
                continue
            if name.lower() in lowered:
                return entity
        return None

    def _tokenize(self, text: str) -> set[str]:
        tokens: set[str] = set()
        for m in TOKEN_RE.finditer(text):
            token = m.group(0).strip().lower()
            if len(token) <= 1:
                continue
            tokens.add(token)
            if self._contains_cjk(token) and len(token) > 2:
                for i in range(len(token) - 1):
                    tokens.add(token[i : i + 2])
        return tokens

    def _score(self, query_tokens: set[str], document: str) -> float:
        if not query_tokens:
            return 0.0
        doc_tokens = self._tokenize(document)
        overlap = query_tokens.intersection(doc_tokens)
        if not overlap:
            return 0.0
        return len(overlap) / max(len(query_tokens), 1)

    def _contains_cjk(self, token: str) -> bool:
        return any("\u4e00" <= ch <= "\u9fff" for ch in token)

    def _merge_entities(
        self,
        primary: list[dict[str, Any]],
        secondary: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        merged: dict[str, dict[str, Any]] = {}
        for row in primary:
            eid = str(row.get("id", "")).strip()
            if eid:
                merged[eid] = row
        for row in secondary:
            eid = str(row.get("id", "")).strip()
            if not eid:
                continue
            merged[eid] = row
        return list(merged.values())

    def _build_image_url(self, raw: dict[str, Any], title: str, source_file: str) -> str | None:
        image_ref = self._extract_image_ref(raw)
        if not image_ref:
            return None
        image_id = self._register_image(
            raw,
            title=title,
            source_file=source_file,
            target_registry=self._images,
            image_ref_override=image_ref,
        )
        if not image_id:
            return None
        return f"/api/v1/image/file/{quote(image_id, safe='')}"

    def _register_image(
        self,
        raw: dict[str, Any],
        title: str,
        source_file: str,
        target_registry: dict[str, dict[str, Any]],
        image_ref_override: str | None = None,
    ) -> str | None:
        image_ref = image_ref_override or self._extract_image_ref(raw)
        if not image_ref:
            return None
        image_id = self._make_image_id(image_ref)
        if image_id in target_registry:
            return image_id
        target_registry[image_id] = {
            "image_id": image_id,
            "title": title,
            "source": source_file,
            "ref": image_ref,
            "kind": "remote" if image_ref.lower().startswith(("http://", "https://")) else "local",
            "url": f"/api/v1/image/file/{quote(image_id, safe='')}",
            "object_keys": self._build_object_keys(image_id, image_ref),
            "bucket": settings.minio_bucket,
        }
        return image_id

    def _make_image_id(self, image_ref: str) -> str:
        digest = hashlib.sha1(image_ref.encode("utf-8")).hexdigest()
        return f"img_{digest[:20]}"

    def _build_object_keys(self, image_id: str, image_ref: str) -> dict[str, str]:
        suffix = Path(image_ref).suffix.lower()
        if not suffix or suffix not in IMAGE_EXTS:
            suffix = ".jpg"
        prefix = f"astro/images/{image_id[:2]}/{image_id}"
        return {
            "original": f"{prefix}/original{suffix}",
            "thumb_1024": f"{prefix}/thumb_1024.webp",
            "thumb_512": f"{prefix}/thumb_512.webp",
            "thumb_256": f"{prefix}/thumb_256.webp",
        }

    def _connect(self) -> sqlite3.Connection:
        return get_sqlite_connection(self._db_path)

    def _init_db(self) -> None:
        conn = self._connect()
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS image_assets (
                    image_id TEXT PRIMARY KEY,
                    title TEXT NOT NULL,
                    source TEXT NOT NULL,
                    ref TEXT NOT NULL,
                    kind TEXT NOT NULL,
                    url TEXT NOT NULL,
                    object_keys_json TEXT NOT NULL,
                    bucket TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS knowledge_chunks (
                    entity_id TEXT PRIMARY KEY,
                    title TEXT NOT NULL,
                    description TEXT NOT NULL,
                    category TEXT NOT NULL,
                    source_file TEXT NOT NULL,
                    raw_json TEXT NOT NULL
                )
                """
            )
            # 索引：加速按 category 查询和 source_file 过滤
            conn.execute("CREATE INDEX IF NOT EXISTS idx_knowledge_chunks_category ON knowledge_chunks(category)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_knowledge_chunks_source ON knowledge_chunks(source_file)")
            # 索引：加速 image_assets 的 title 模糊查询和 kind 过滤
            conn.execute("CREATE INDEX IF NOT EXISTS idx_image_assets_title ON image_assets(title)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_image_assets_kind ON image_assets(kind)")
            conn.commit()
        finally:
            conn.close()

    def _persist_images_metadata(self, images: dict[str, dict[str, Any]]) -> None:
        if not images:
            return
        conn = self._connect()
        try:
            rows = [
                (
                    item["image_id"],
                    item.get("title", ""),
                    item.get("source", ""),
                    item.get("ref", ""),
                    item.get("kind", ""),
                    item.get("url", ""),
                    self._json_dump(item.get("object_keys", {})),
                    item.get("bucket", settings.minio_bucket),
                )
                for item in images.values()
            ]
            conn.executemany(
                """
                INSERT INTO image_assets (image_id, title, source, ref, kind, url, object_keys_json, bucket)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(image_id) DO UPDATE SET
                    title = excluded.title,
                    source = excluded.source,
                    ref = excluded.ref,
                    kind = excluded.kind,
                    url = excluded.url,
                    object_keys_json = excluded.object_keys_json,
                    bucket = excluded.bucket
                """,
                rows,
            )
            conn.commit()
        finally:
            conn.close()

    def _load_persisted_extra_entities(self) -> list[dict[str, Any]]:
        conn = self._connect()
        try:
            rows = conn.execute(
                """
                SELECT entity_id, title, description, category, source_file, raw_json
                FROM knowledge_chunks
                ORDER BY entity_id ASC
                """
            ).fetchall()
        finally:
            conn.close()
        entities: list[dict[str, Any]] = []
        for row in rows:
            payload = dict(row)
            entities.append(
                {
                    "id": payload["entity_id"],
                    "name": payload["title"],
                    "description": payload["description"],
                    "category": payload["category"],
                    "source_file": payload["source_file"],
                    "raw": self._json_load(payload.get("raw_json", "{}")),
                }
            )
        return entities

    def _persist_extra_entities(self, entities: list[dict[str, Any]]) -> None:
        if not entities:
            return
        rows = [
            (
                str(e.get("id", "")),
                str(e.get("name", "")),
                str(e.get("description", "")),
                str(e.get("category", "")),
                str(e.get("source_file", "")),
                self._json_dump(e.get("raw", {})),
            )
            for e in entities
            if str(e.get("id", "")).strip()
        ]
        if not rows:
            return
        conn = self._connect()
        try:
            conn.executemany(
                """
                INSERT INTO knowledge_chunks (entity_id, title, description, category, source_file, raw_json)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(entity_id) DO UPDATE SET
                    title = excluded.title,
                    description = excluded.description,
                    category = excluded.category,
                    source_file = excluded.source_file,
                    raw_json = excluded.raw_json
                """,
                rows,
            )
            conn.commit()
        finally:
            conn.close()

    def _clear_persisted_extra_entities(self) -> None:
        conn = self._connect()
        try:
            conn.execute("DELETE FROM knowledge_chunks")
            conn.commit()
        finally:
            conn.close()

    def _rewrite_persisted_extra_entities(self, entities: list[dict[str, Any]]) -> None:
        conn = self._connect()
        try:
            conn.execute("DELETE FROM knowledge_chunks")
            rows = [
                (
                    str(e.get("id", "")),
                    str(e.get("name", "")),
                    str(e.get("description", "")),
                    str(e.get("category", "")),
                    str(e.get("source_file", "")),
                    self._json_dump(e.get("raw", {})),
                )
                for e in entities
                if str(e.get("id", "")).strip()
            ]
            if rows:
                conn.executemany(
                    """
                    INSERT INTO knowledge_chunks (entity_id, title, description, category, source_file, raw_json)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    rows,
                )
            conn.commit()
        finally:
            conn.close()

    def _json_dump(self, obj: Any) -> str:
        import json

        return json.dumps(obj, ensure_ascii=False)

    def _json_load(self, raw: str) -> Any:
        import json

        try:
            return json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            return {}

    def _extract_image_ref(self, raw: dict[str, Any]) -> str | None:
        candidate_keys = (
            "image_url",
            "image",
            "image_path",
            "image link",
            "image url",
            "img_url",
            "img_path",
            "thumbnail_url",
            "thumbnail",
            "thumb",
            "photo",
            "photo_url",
            "file_path",
            "filepath",
            "path",
            "url",
            "图片",
            "图片地址",
            "图片路径",
            "图像",
            "图像地址",
            "图像路径",
            "缩略图",
            "封面",
            "链接",
        )
        for key in candidate_keys:
            value = str(raw.get(key, "")).strip()
            if not value:
                continue
            if value.lower().startswith(("http://", "https://")):
                return value
            resolved = self._resolve_local_path(value)
            if resolved:
                return resolved
        return None

    def _resolve_local_path(self, value: str) -> str | None:
        if value.lower().startswith("file://"):
            value = value[7:]
        value_path = Path(value)
        candidates: list[Path] = []
        if value_path.is_absolute():
            candidates.append(value_path)
        if self._source_root:
            root = Path(self._source_root)
            base = root.parent if root.is_file() else root
            candidates.append(base / value)
            candidates.append(base.parent / value)
        image_base_dirs = [x.strip() for x in settings.image_base_dirs.split(",") if x.strip()]
        for base_dir in image_base_dirs:
            base = Path(base_dir)
            candidates.append(base / value)
            candidates.append(base.parent / value)
            if value_path.parts:
                try:
                    candidates.append(base / Path(*value_path.parts[-4:]))
                except Exception:  # noqa: BLE001
                    pass
        candidates.append(value_path)
        for candidate in candidates:
            if not candidate.exists() or not candidate.is_file():
                continue
            if candidate.suffix.lower() not in IMAGE_EXTS:
                continue
            try:
                return str(candidate.resolve())
            except OSError:
                return str(candidate)
        return None
